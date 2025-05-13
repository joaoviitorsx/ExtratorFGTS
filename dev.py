import warnings
import pdfplumber
import re
import json
import os
import webbrowser
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from tkinter import *
from tkinter import ttk, filedialog, messagebox

dados_extraidos = []

def extrair_dados_fgts_pdfplumber(caminho_pdf):
    import pdfplumber
    import re
    import warnings
    
    warnings.filterwarnings('ignore')
    dados_por_competencia = {}
    competencia_atual = None
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if not texto:
                    continue

                match_comp = re.search(r"(?i)\bcompet[êeéè]ncia\:?\s*(\d{2}/\d{4})", texto)
                if match_comp:
                    competencia_atual = match_comp.group(1)

                if not competencia_atual:
                    continue

                blocos_encontrados = re.findall(r"(?:^|\n|\s)Empr\.\:?\s*\d+.*?(?=(?:\n|\s)Empr\.\:|$)", texto, re.DOTALL)
                
                if not blocos_encontrados:
                    blocos = re.split(r"\n(?=Empr\.\:\s*\d+)", texto)
                else:
                    blocos = blocos_encontrados

                for bloco in blocos:
                    if not re.search(r"Empr\.\:?", bloco):
                        continue
                    
                    match_dados = re.search(
                        r"Empr\.\:?\s*(\d+)\s*([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÜÇÑ\s\-\.]+?)(?:Situação|CPF)\:"  
                        r".*?CPF\:\s*([\d\.\-]+)"            
                        r".*?Adm\:\s*(\d{2}/\d{2}/\d{4}|\d{2}/\d{2}/\d{2}|\d{2}/\d{4})",
                        bloco, 
                        flags=re.DOTALL
                    )
                    
                    if not match_dados:
                        match_dados = re.search(
                            r"Empr\.\:?\s*(\d+)([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÜÇÑ\s\-\.]+?)(?=Situação\:|CPF\:)"
                            r".*?CPF\:\s*([\d\.\-]+)"                           
                            r".*?Adm\:\s*(\d{2}/\d{2}/\d{4}|\d{2}/\d{2}/\d{2}|\d{2}/\d{4})",
                            bloco, 
                            flags=re.DOTALL
                        )
                    
                    if not match_dados:
                        match_matricula = re.search(r"Empr\.\:?\s*(\d+)", bloco)
                        match_cpf = re.search(r"CPF\:\s*([\d\.\-]+)", bloco)
                        match_adm = re.search(r"Adm\:\s*(\d{2}/\d{2}/\d{4}|\d{2}/\d{2}/\d{2}|\d{2}/\d{4})", bloco)
                        
                        if match_matricula and match_cpf and match_adm:
                            matricula = match_matricula.group(1)
                            
                            nome = ""
                            inicio_texto = bloco[match_matricula.end():]
                            match_fim = re.search(r"\s+(?:Situação|CPF)\:", inicio_texto)
                            
                            if match_fim:
                                nome = inicio_texto[:match_fim.start()].strip()
                            
                            if not nome or len(nome) < 3:
                                palavras = bloco.split()
                                for i in range(len(palavras)):
                                    if palavras[i].startswith("Empr.") and i+2 < len(palavras):
                                        possivel_nome = []
                                        for j in range(i+2, len(palavras)):
                                            if "CPF:" in palavras[j] or "Situação:" in palavras[j]:
                                                break
                                            possivel_nome.append(palavras[j])
                                        nome = " ".join(possivel_nome).strip()
                                        break
                            
                            class MockMatch:
                                def groups(self):
                                    return (matricula.strip(), nome, 
                                            match_cpf.group(1), match_adm.group(1))
                            
                            match_dados = MockMatch()

                    if not match_dados:
                        continue

                    matricula, nome, cpf, admissao = match_dados.groups()
                    
                    nome = re.sub(r"\s+", " ", nome.strip())
                    nome = nome.title()
                    
                    match_fgts = re.search(
                        r"Base\s*FGTS\:?\s*([\d\.,]+)[\s\n]*Valor\s*FGTS\:?\s*([\d\.,]+)", 
                        bloco
                    )
                    
                    if not match_fgts:
                        match_fgts = re.search(
                            r"Base\s*FGTS\:?\s*([\d\.,]+).*?Valor\s*FGTS\:?\s*([\d\.,]+)",
                            bloco,
                            flags=re.DOTALL
                        )
                    
                    if not match_fgts:
                        linhas = bloco.split('\n')
                        for i, linha in enumerate(linhas):
                            if "Base FGTS" in linha:
                                match_base = re.search(r"Base\s*FGTS\:?\s*([\d\.,]+)", linha)
                                for j in range(i, min(i+5, len(linhas))):
                                    match_valor = re.search(r"Valor\s*FGTS\:?\s*([\d\.,]+)", linhas[j])
                                    if match_valor and match_base:
                                        class FGTSMatch:
                                            def groups(self):
                                                return match_base.group(1), match_valor.group(1)
                                        match_fgts = FGTSMatch()
                                        break
                                if match_fgts:
                                    break
                                    
                    if not match_fgts:
                        match_base = re.search(r"Base\s*FGTS\:?\s*([\d\.,]+)", bloco)
                        match_valor = re.search(r"Valor\s*FGTS\:?\s*([\d\.,]+)", bloco)
                        
                        if match_base and match_valor:
                            class FGTSMatch:
                                def groups(self):
                                    return match_base.group(1), match_valor.group(1)
                            match_fgts = FGTSMatch()
                                    
                    if not match_fgts:
                        continue

                    base_fgts, valor_fgts = match_fgts.groups()
                    
                    try:
                        base_fgts_num = base_fgts.replace(".", "").replace(",", ".")
                        valor_fgts_num = valor_fgts.replace(".", "").replace(",", ".")
                        
                        _ = float(base_fgts_num)
                        _ = float(valor_fgts_num)
                    except ValueError:
                        continue

                    registro = {
                        "Matricula": matricula.strip(),
                        "Empregado": nome,
                        "CPF": cpf.strip(),
                        "Admissao": admissao.strip(),
                        "Base FGTS": base_fgts_num,
                        "Valor FGTS": valor_fgts_num
                    }

                    if all(registro.values()):
                        dados_por_competencia.setdefault(competencia_atual, []).append(registro)

        return dados_por_competencia
        
    except Exception as e:
        print(f"Erro ao processar o PDF: {str(e)}")
        return {}
    
def extrair_dados_folha_pagamento_pdfplumber(caminho_pdf):
    import pdfplumber
    import re
    import warnings
    
    warnings.filterwarnings('ignore')
    dados_por_competencia = {}
    
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto_completo = ""
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if not texto:
                    continue
                texto_completo += texto + "\n"
            
            # Extrair mês/ano de referência
            match_mes_ano = re.search(r"Mês/Ano:\s*(\d{2}/\d{4})", texto_completo)
            competencia = match_mes_ano.group(1) if match_mes_ano else "00/0000"
            
            # Identificar funcionários e suas informações
            # Padrão: 6 dígitos seguidos do nome do funcionário
            padrao_funcionario = r"(\d{6})\s+([A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÜÇÑ\s]+?)(?:\nCargo:|Continuação)"
            funcionarios = re.finditer(padrao_funcionario, texto_completo)
            
            for match in funcionarios:
                matricula = match.group(1).strip()
                nome = match.group(2).strip()
                
                # Encontrar início do bloco do funcionário
                inicio = match.start()
                
                # Encontrar final do bloco (próximo funcionário ou fim do texto)
                proximo = re.search(r"\d{6}\s+[A-ZÁÀÂÃÉÈÊÍÏÓÔÕÖÚÜÇÑ\s]+?\n", texto_completo[match.end():])
                fim = match.end() + proximo.start() if proximo else len(texto_completo)
                
                bloco = texto_completo[inicio:fim]
                
                # Extrair valor FGTS
                match_fgts = re.search(r"FGTS:\s+([\d\.,]+)", bloco)
                valor_fgts = "0.00"
                if match_fgts:
                    valor_fgts = match_fgts.group(1).replace(".", "").replace(",", ".")
                
                # Extrair base de cálculo FGTS
                match_base_fgts = re.search(r"BC-FGTS:\s+([\d\.,]+)", bloco)
                base_fgts = "0.00"
                if match_base_fgts:
                    base_fgts = match_base_fgts.group(1).replace(".", "").replace(",", ".")
                
                # Melhorar a extração da data de admissão - método 1
                match_admissao = re.search(r"Admissão\s+(\d{2}/\d{2}/\d{4})", bloco)
                
                # Método 2: Procurar padrão de linha após "Admissão"
                if not match_admissao:
                    linhas = bloco.split('\n')
                    for i, linha in enumerate(linhas):
                        if "Admissão" in linha and i+1 < len(linhas):
                            # Procurar data no início da próxima linha
                            match_data = re.search(r"^(\d{2}/\d{2}/\d{4})", linhas[i+1].strip())
                            if match_data:
                                match_admissao = match_data
                                break
                
                # Método 3: Procurar por data em formato DD/MM/AAAA após "Data: Assinatura:"
                if not match_admissao:
                    match_data_assinatura = re.search(r"Data:\s*Assinatura:.*?(\d{2}/\d{2}/\d{4})", bloco, re.DOTALL)
                    if match_data_assinatura:
                        match_admissao = match_data_assinatura
                
                data_admissao = match_admissao.group(1) if match_admissao else ""
                
                # Criar registro no formato esperado
                registro = {
                    "Matricula": matricula,
                    "Empregado": nome,
                    "CPF": "",  # CPF não disponível neste layout
                    "Admissao": data_admissao,
                    "Base FGTS": base_fgts,
                    "Valor FGTS": valor_fgts
                }
                
                # Adicionar ao dicionário apenas se base FGTS ou valor FGTS forem diferentes de zero
                if float(base_fgts) > 0 or float(valor_fgts) > 0:
                    dados_por_competencia.setdefault(competencia, []).append(registro)
            
        return dados_por_competencia
        
    except Exception as e:
        print(f"Erro ao processar o PDF de folha de pagamento: {str(e)}")
        return {}
    
def atualizar_visualizacao(registros_por_competencia):
    global dados_extraidos
    dados_extraidos = registros_por_competencia

    total_registros = sum(len(lista) for lista in registros_por_competencia.values())
    total_competencias = len(registros_por_competencia)

    text_widget.delete(1.0, END)
    text_widget.insert(END, json.dumps(registros_por_competencia, indent=4, ensure_ascii=False))

    status_var.set(
        f"{total_registros} empregados extraídos em {total_competencias} competência(s)."
    )

def processar_arquivo(caminho):
    registros = extrair_dados_fgts_pdfplumber(caminho)
    
    if not registros:
        registros = extrair_dados_folha_pagamento_pdfplumber(caminho)
        
    atualizar_visualizacao(registros)

def processar_pasta():
    pasta = filedialog.askdirectory()
    if not pasta:
        return
    registros_por_competencia = {}

    for arquivo in os.listdir(pasta):
        if arquivo.lower().endswith(".pdf"):
            caminho_completo = os.path.join(pasta, arquivo)
            dados_pdf = extrair_dados_fgts_pdfplumber(caminho_completo)

            for competencia, registros in dados_pdf.items():
                registros_por_competencia.setdefault(competencia, []).extend(registros)

    atualizar_visualizacao(registros_por_competencia)

def escolher_pdf():
    caminho = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
    if caminho:
        processar_arquivo(caminho)

def salvar_planilha_formatada():
    if not dados_extraidos:
        messagebox.showwarning("Aviso", "Nenhum dado foi extraído para salvar.")
        return

    caminho_saida = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Planilhas Excel", "*.xlsx")],
        title="Salvar planilha formatada"
    )
    if not caminho_saida:
        return

    with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
        competencias_ordenadas = sorted(
            dados_extraidos.keys(),
            key=lambda x: pd.to_datetime(f"01/{x}", dayfirst=True)
        )

        for competencia in competencias_ordenadas:
            registros = dados_extraidos[competencia]
            dados_formatados = []
            for item in registros:
                item_copia = item.copy()
                try:
                    item_copia["Base FGTS"] = f"{float(item_copia['Base FGTS']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    item_copia["Valor FGTS"] = f"{float(item_copia['Valor FGTS']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                except (ValueError, TypeError):
                    item_copia["Base FGTS"] = item_copia.get("Base FGTS", "")
                    item_copia["Valor FGTS"] = item_copia.get("Valor FGTS", "")
                dados_formatados.append(item_copia)

            df = pd.DataFrame(dados_formatados, columns=["Matricula", "Empregado", "CPF", "Admissao", "Base FGTS", "Valor FGTS"])
            aba_nome = competencia.replace("/", "_")
            df.to_excel(writer, sheet_name=aba_nome, index=False)

    wb = load_workbook(caminho_saida)
    fonte_arial = Font(name='Arial', size=10)
    for aba in wb.worksheets:
        for row in aba.iter_rows(min_row=2):
            for i, cell in enumerate(row):
                cell.font = fonte_arial
                if i == 0:
                    cell.number_format = "@"
                if i in (4, 5):
                    cell.number_format = '#.##0,00'
                else:
                    cell.number_format = "@"
    wb.save(caminho_saida)

    if messagebox.askyesno("Planilha salva", "Deseja abrir a planilha agora?"):
        webbrowser.open(f"file://{os.path.abspath(caminho_saida)}")

def visualizar_texto_bruto_pdf():
    """Permite visualizar o texto bruto extraído pelo pdfplumber de um PDF selecionado."""
    caminho = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
    if not caminho:
        return
    
    debug_window = Toplevel(root)
    debug_window.title("Debug - Texto Extraído pelo PDFPlumber")
    debug_window.geometry("900x700")
    
    frame_controle = Frame(debug_window)
    frame_controle.pack(fill=X, padx=10, pady=5)
    
    pagina_var = StringVar(value="Página: -")
    Label(frame_controle, textvariable=pagina_var).pack(side=LEFT, padx=5)
    
    frame_texto = Frame(debug_window)
    frame_texto.pack(fill=BOTH, expand=True, padx=10, pady=5)
    
    scrollbar_y = Scrollbar(frame_texto)
    scrollbar_y.pack(side=RIGHT, fill=Y)
    
    scrollbar_x = Scrollbar(frame_texto, orient=HORIZONTAL)
    scrollbar_x.pack(side=BOTTOM, fill=X)
    
    debug_text = Text(frame_texto, wrap="none", yscrollcommand=scrollbar_y.set, 
                      xscrollcommand=scrollbar_x.set, font=("Courier New", 10))
    debug_text.pack(side=LEFT, fill=BOTH, expand=True)
    
    scrollbar_y.config(command=debug_text.yview)
    scrollbar_x.config(command=debug_text.xview)
    
    paginas_texto = []
    pagina_atual = [0]
    
    def carregar_pdf():
        try:
            paginas_texto.clear()
            debug_text.delete(1.0, END)
            
            with pdfplumber.open(caminho) as pdf:
                for i, pagina in enumerate(pdf.pages):
                    texto = pagina.extract_text() or f"[Página {i+1}: Sem texto extraível]"
                    paginas_texto.append(texto)
            
            if paginas_texto:
                pagina_atual[0] = 0
                mostrar_pagina()
                return True
        except Exception as e:
            debug_text.delete(1.0, END)
            debug_text.insert(END, f"Erro ao processar o PDF:\n{str(e)}")
            return False
    
    def mostrar_pagina():
        if not paginas_texto:
            return
        
        debug_text.delete(1.0, END)
        debug_text.insert(END, paginas_texto[pagina_atual[0]])
        pagina_var.set(f"Página: {pagina_atual[0] + 1} de {len(paginas_texto)}")
    
    def proxima_pagina():
        if not paginas_texto or pagina_atual[0] >= len(paginas_texto) - 1:
            return
        pagina_atual[0] += 1
        mostrar_pagina()
    
    def pagina_anterior():
        if not paginas_texto or pagina_atual[0] <= 0:
            return
        pagina_atual[0] -= 1
        mostrar_pagina()
    
    def mostrar_todas_paginas():
        if not paginas_texto:
            return
        
        debug_text.delete(1.0, END)
        for i, texto in enumerate(paginas_texto):
            debug_text.insert(END, f"\n\n{'=' * 40}\nPÁGINA {i + 1}\n{'=' * 40}\n\n")
            debug_text.insert(END, texto)
        
        pagina_var.set(f"Mostrando todas as {len(paginas_texto)} páginas")
    
    def salvar_texto():
        if not paginas_texto:
            messagebox.showwarning("Aviso", "Não há texto para salvar.")
            return
        
        arquivo_saida = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Arquivos de Texto", "*.txt")],
            title="Salvar Texto Extraído"
        )
        
        if not arquivo_saida:
            return
        
        try:
            with open(arquivo_saida, "w", encoding="utf-8") as f:
                for i, texto in enumerate(paginas_texto):
                    f.write(f"\n\n{'=' * 40}\nPÁGINA {i + 1}\n{'=' * 40}\n\n")
                    f.write(texto)
            
            messagebox.showinfo("Sucesso", f"Texto salvo com sucesso em:\n{arquivo_saida}")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar o arquivo:\n{str(e)}")
    
    btn_frame = Frame(frame_controle)
    btn_frame.pack(side=RIGHT)
    
    ttk.Button(btn_frame, text="Página Anterior", command=pagina_anterior).pack(side=LEFT, padx=2)
    ttk.Button(btn_frame, text="Próxima Página", command=proxima_pagina).pack(side=LEFT, padx=2)
    ttk.Button(btn_frame, text="Todas as Páginas", command=mostrar_todas_paginas).pack(side=LEFT, padx=2)
    ttk.Button(btn_frame, text="Salvar Texto", command=salvar_texto).pack(side=LEFT, padx=2)
    
    if carregar_pdf():
        debug_window.focus_set()
    else:
        messagebox.showerror("Erro", "Não foi possível processar o PDF selecionado.")
        debug_window.destroy()

root = Tk()
root.title("Extrator de FGTS de PDFs - Assertivus")
root.geometry("1100x700")

style = ttk.Style()
style.configure("TButton", font=("Segoe UI", 10, "bold"))

frame_botoes = ttk.LabelFrame(root, text="Ações")
frame_botoes.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

btn_pdf = ttk.Button(frame_botoes, text="Selecionar PDF Único", command=escolher_pdf)
btn_pdf.grid(row=0, column=0, padx=5, pady=5)

btn_pasta = ttk.Button(frame_botoes, text="Selecionar Pasta de PDFs", command=processar_pasta)
btn_pasta.grid(row=0, column=1, padx=5, pady=5)

btn_salvar = ttk.Button(frame_botoes, text="Gerar Planilha Formatada", command=salvar_planilha_formatada)
btn_salvar.grid(row=0, column=2, padx=5, pady=5)

btn_debug = ttk.Button(frame_botoes, text="Debug PDF", command=visualizar_texto_bruto_pdf)
btn_debug.grid(row=0, column=3, padx=5, pady=5)

frame_texto = Frame(root)
frame_texto.grid(row=1, column=0, padx=10, pady=(0,10), sticky="nsew")
root.grid_rowconfigure(1, weight=1)
root.grid_columnconfigure(0, weight=1)

scrollbar = Scrollbar(frame_texto)
scrollbar.pack(side=RIGHT, fill=Y)

text_widget = Text(frame_texto, wrap="word", yscrollcommand=scrollbar.set, font=("Courier New", 10))
text_widget.pack(side=LEFT, fill=BOTH, expand=True)
scrollbar.config(command=text_widget.yview)

status_var = StringVar()
status_var.set("Nenhum dado carregado.")
status_bar = Label(root, textvariable=status_var, bd=1, relief=SUNKEN, anchor="w")
status_bar.grid(row=2, column=0, sticky="ew")

root.mainloop()