import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def gerar_planilha_fgts(dados, caminho_saida):
    colunas = ["Matricula", "Empregado", "Admissao", "CPF", "Base FGTS", "Valor FGTS"]
    df = pd.DataFrame(dados, columns=colunas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados FGTS"

    arial_font = Font(name="Arial", size=11)
    header_font = Font(name="Arial", bold=True, size=11)

    for col_idx, col_name in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font

    for row_idx, linha in enumerate(dados, start=2):
        for col_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.font = arial_font

    for col_idx, col_name in enumerate(colunas, start=1):
        coluna_letra = get_column_letter(col_idx)
        valores_coluna = [str(ws.cell(row=row, column=col_idx).value) for row in range(1, len(dados)+2)]
        largura_maxima = max(len(str(col_name)), max(len(valor) for valor in valores_coluna))
        ws.column_dimensions[coluna_letra].width = largura_maxima + 2

    wb.save(caminho_saida)
    print(f"Planilha gerada com sucesso: {caminho_saida}")
